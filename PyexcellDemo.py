import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By

workbook = openpyxl.Workbook
sheet = workbook.active
driver = webdriver.Chrome(executable_path="Driver/chromedriver.exe")
driver.get("https://rahulshettyacademy.com/AutomationPractice/")
number_of_row = driver.find_elements(By.XPATH,"//div[@class='tableFixHead']//tbody/tr")
for item1 in number_of_row:
    sheet.cell(1,1).value = driver.find_element(By.XPATH,"//div[@class='tableFixHead']//tbody/tr[1]/td[1]")
    sheet.cell(1,1).value = driver.find_element(By.XPATH,"//div[@class='tableFixHead']//tbody/tr[1]/td[1]")
    sheet.cell(1,1).value = driver.find_element(By.XPATH,"//div[@class='tableFixHead']//tbody/tr[1]/td[1]")

sheet.save("output.xlsx")